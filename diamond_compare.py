import os
import pandas as pd
import json
import datetime
from itertools import combinations
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from logger import get_logger

load_dotenv()
log = get_logger("comparator")

# ── Directories ────────────────────────────────────────────────
DIAMOND_FILES_DIR       = os.path.join(os.getcwd(), "diamond_files")
COMPARE_DIR             = os.path.join(os.getcwd(), "compare")
FILE_STATUS_JSON        = os.path.join(DIAMOND_FILES_DIR, "file_status.json")
COMPARE_CSV              = os.path.join(COMPARE_DIR, "compare.csv")
COMPARE_JSON             = os.path.join(COMPARE_DIR, "compare.json")
COMPARE_PARTIAL_JSON     = os.path.join(COMPARE_DIR, "compare_partial.json")
COMPARE_ALL_JSON         = os.path.join(COMPARE_DIR, "compare_all.json")
COMPARE_BR_VS_LG_JSON    = os.path.join(COMPARE_DIR, "compare_br_vs_lg.json")   # ← NEW
DIAMOND_STATUS_JSON      = os.path.join(COMPARE_DIR, "diamond_status.json")
MAX_AGE_DAYS             = 7

os.makedirs(COMPARE_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════
# DATA LOADERS
# ══════════════════════════════════════════════════════════════

def load_vendor(csv_path, cert_col, price_col, name):
    log.debug(f"Loading vendor '{name}' from {csv_path}")
    df = pd.read_csv(csv_path, low_memory=False, on_bad_lines="skip")
    df["certificate_number"] = df[cert_col].astype(str).str.strip()
    df = df[["certificate_number", price_col]].copy()
    df = df.rename(columns={price_col: f"{name} Price USD"})
    df[f"{name} Price USD"] = pd.to_numeric(df[f"{name} Price USD"], errors="coerce")
    log.info(f"Loaded '{name}': {len(df):,} rows")
    return df


PC_EXTRA_COLS = {
    "SHAPE":   "shape",
    "COLOR":   "color",
    "CLARITY": "clarity",
    "CUT":     "cut",
    "CARAT":   "caratWeight",
}

def load_PC(csv_path):
    log.debug(f"Loading PC from {csv_path}")
    df = pd.read_csv(csv_path, low_memory=False, on_bad_lines="skip")
    df = df.rename(columns={"certificateNumber": "certificate_number", "finalPriceUsd": "PC Price USD"})
    df["certificate_number"] = df["certificate_number"].astype(str).str.strip()
    df["PC Price USD"] = pd.to_numeric(df["PC Price USD"], errors="coerce")

    for src_col, dest_col in PC_EXTRA_COLS.items():
        for variant in [src_col, src_col.lower(), src_col.title()]:
            if variant in df.columns:
                df[dest_col] = df[variant].astype(str).str.strip().str.title()
                log.debug(f"  PC extra col: {variant} → {dest_col}")
                break

    keep = ["certificate_number", "PC Price USD"] + [
        dest for dest in PC_EXTRA_COLS.values() if dest in df.columns
    ]
    log.info(f"Loaded PC: {len(df):,} rows | extra cols: {[c for c in keep if c not in ('certificate_number','PC Price USD')]}")
    return df[keep]


# ══════════════════════════════════════════════════════════════
# NEW — BUILD compare_br_vs_lg.json
#   Diamonds present in BOTH Brilliance AND Loose-Grown CSVs,
#   regardless of whether they exist in PC.
#   Includes their raw prices, any PC price if found, and
#   % diff columns so the dashboard table can render them.
# ══════════════════════════════════════════════════════════════

LG_EXTRA_COLS = {
    "shape":   "shape",
    "carat":   "caratWeight",
    "cut":     "cut",
    "color":   "color",
    "clarity": "clarity",
}

def load_lg_with_attrs(csv_path: str, cert_col: str, price_col: str) -> pd.DataFrame:
    """Load LG CSV keeping certificate_number, price, and diamond attribute columns."""
    log.debug(f"Loading LG with attrs from {csv_path}")
    df = pd.read_csv(csv_path, low_memory=False, on_bad_lines="skip")
    log.info(f"  LG CSV columns: {list(df.columns)}")

    # Normalise column names to lowercase for lookup
    col_map = {c.lower().strip(): c for c in df.columns}

    df["certificate_number"] = df[cert_col].astype(str).str.strip()
    df = df.rename(columns={price_col: "loose-grown Price USD"})
    df["loose-grown Price USD"] = pd.to_numeric(df["loose-grown Price USD"], errors="coerce")
    keep = ["certificate_number", "loose-grown Price USD"]

    for src_col_lower, dest_col in LG_EXTRA_COLS.items():
        actual = col_map.get(src_col_lower)
        if actual and actual in df.columns:
            if dest_col == "caratWeight":
                df[dest_col] = pd.to_numeric(df[actual], errors="coerce")
            else:
                df[dest_col] = df[actual].astype(str).str.strip().str.title()
            keep.append(dest_col)
            log.info(f"  LG extra col: '{actual}' → '{dest_col}'")
        else:
            log.warning(f"  LG CSV missing column for '{src_col_lower}' (looked in: {list(col_map.keys())})")

    log.info(f"Loaded LG with attrs: {len(df):,} rows | kept: {keep}")
    return df[[c for c in keep if c in df.columns]]


def build_br_vs_lg(loaded: dict, discounts: dict) -> int:
    """
    Build compare_br_vs_lg.json:
      - Inner join of Brilliance ∩ Loose-Grown on certificate_number
      - Uses LG CSV directly (with attrs) so shape/carat/cut/color/clarity are populated
      - Renames certificate_number → diamond_id, fixes _x/_y merge cols
      - NO PC data — only BR and LG prices + BR -30% vs LG comparison
      - Returns count of records saved
    """
    log.info("=" * 50)
    log.info("BUILDING compare_br_vs_lg.json (Brilliance ∩ LG)")

    # Re-load LG with attribute columns (the loaded dict only has cert+price)
    from dotenv import load_dotenv
    load_dotenv()
    lg_csv   = os.path.join(DIAMOND_FILES_DIR, get_env("LOOSE_GROWN_CSV", "loosegrowndiamond.csv"))
    lg_cert  = get_env("LOOSE_GROWN_CERT_COL", "sku")
    lg_price = get_env("LOOSE_GROWN_PRICE_COL", "price")
    lg_df    = load_lg_with_attrs(lg_csv, lg_cert, lg_price)

    br_df = loaded.get("brilliance")
    pc_df = loaded.get("PC")

    if br_df is None:
        log.warning("brilliance not loaded — skipping br_vs_lg build")
        return 0

    # Inner join: only certs present in BOTH BR and LG
    merged = pd.merge(lg_df, br_df, on="certificate_number", how="inner")
    log.info(f"  LG rows: {len(lg_df):,}  |  BR rows: {len(br_df):,}  |  Inner join: {len(merged):,}")

    # Keep only rows where BOTH LG and BR prices are present
    before = len(merged)
    merged = merged.dropna(subset=["loose-grown Price USD", "brilliance Price USD"])
    merged = merged[merged["loose-grown Price USD"] > 0]
    merged = merged[merged["brilliance Price USD"] > 0]
    log.info(f"  After price filter: {len(merged):,} rows (dropped {before - len(merged):,} with null/zero prices)")

    if merged.empty:
        log.warning("No overlapping cert IDs between Brilliance and LG — br_vs_lg will be empty")
        with open(COMPARE_BR_VS_LG_JSON, "w", encoding="utf-8") as f:
            json.dump([], f)
        return 0

    # ── Fix _x / _y duplicate columns from merge ──────────────
    # LG has attrs (shape_x, color_x ...), BR has none → drop _y cols, rename _x → clean
    for col in list(merged.columns):
        if col.endswith("_y"):
            merged = merged.drop(columns=[col])
        elif col.endswith("_x"):
            merged = merged.rename(columns={col: col[:-2]})

    # ── Rename certificate_number → diamond_id ─────────────────
    merged = merged.rename(columns={"certificate_number": "diamond_id"})

    # ── Add discounted price columns (BR only, skip PC) ────────
    for name, disc in discounts.items():
        if name == "PC":
            continue
        price_col = f"{name} Price USD"
        label     = f"-{round((1 - disc) * 100)}%"
        disc_col  = f"{name} {label} USD"
        if price_col in merged.columns:
            merged[price_col] = pd.to_numeric(merged[price_col], errors="coerce").round(0)
            merged[disc_col]  = (merged[price_col] * disc).round(0)

    # ── Add BR -30% vs LG price comparison ────────────────────
    br_disc      = discounts.get("brilliance", 0.70)
    br_label     = f"-{round((1 - br_disc) * 100)}%"
    br_disc_col  = f"brilliance {br_label} USD"
    lg_price_col = "loose-grown Price USD"
    if br_disc_col in merged.columns and lg_price_col in merged.columns:
        br_vals = pd.to_numeric(merged[br_disc_col], errors="coerce").replace(0, float("nan"))
        lg_vals = pd.to_numeric(merged[lg_price_col], errors="coerce").replace(0, float("nan"))
        merged["BR -30% vs LG USD"] = (br_vals - lg_vals).round(0)
        merged["BR -30% vs LG %"]   = ((br_vals - lg_vals) / lg_vals * 100).round(2)
        log.info(f"Added BR -30% vs LG comparison  ({merged['BR -30% vs LG %'].notna().sum():,} valid rows)")

    # ── Drop all PC-related and helper columns ─────────────────
    pc_cols_to_drop = (
        ["PC Price USD", "loose-grown vs PC %", "brilliance vs PC %",
         "luvansh vs PC %", "vendors_matched", "in_pc",
         "_all_matched", "_some_matched", "_vendor_count"]
        + [c for c in merged.columns if c.startswith("PC ")]
    )
    merged = merged.drop(columns=pc_cols_to_drop, errors="ignore")

    # Clean for JSON
    import math

    records = merged.where(merged.notna(), other=None).to_dict(orient="records")
    clean = []
    for row in records:
        clean_row = {}
        for k, v in row.items():
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                clean_row[k] = None
            else:
                clean_row[k] = v
        clean.append(clean_row)

    with open(COMPARE_BR_VS_LG_JSON, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)

    log.info(f"Saved compare_br_vs_lg.json → {COMPARE_BR_VS_LG_JSON}")
    log.info(f"  Total: {len(clean):,} records | columns: {list(merged.columns)}")
    return len(clean)


# ══════════════════════════════════════════════════════════════
# STEP 1 — BUILD compare files
# ══════════════════════════════════════════════════════════════

def _save_diamond_status(combined: pd.DataFrame, loaded: dict, discounts: dict,
                         common_records: list, partial_records: list, all_records: list,
                         br_vs_lg_count: int = 0):
    import math

    vendor_names = [n for n in loaded.keys() if n != "PC"]
    pc_total     = len(loaded["PC"]) if "PC" in loaded else 0

    vendor_totals = {}
    for name in vendor_names:
        vendor_totals[name] = len(loaded[name]) if name in loaded else 0

    pc_vs = {}
    for name in vendor_names:
        price_col = f"{name} Price USD"
        if price_col not in combined.columns:
            pc_vs[name] = {"matched": 0, "vendor_total": vendor_totals.get(name, 0)}
            continue
        matched   = int(combined[price_col].notna().sum())
        unmatched = int(combined[price_col].isna().sum())
        pc_vs[name] = {
            "matched":          matched,
            "vendor_total":     vendor_totals.get(name, 0),
            "vendor_not_in_pc": vendor_totals.get(name, 0) - matched,
            "match_rate_pct":   round(matched / pc_total * 100, 2) if pc_total else 0,
        }

    status = {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pc_total":     pc_total,
        "vendor_totals": vendor_totals,
        "pc_vs_vendors": pc_vs,
        "dataset_counts": {
            "common":     len(common_records),
            "partial":    len(partial_records),
            "all_pc":     len(all_records),
            "br_vs_lg":   br_vs_lg_count,      # ← NEW
        },
    }

    with open(DIAMOND_STATUS_JSON, "w", encoding="utf-8") as f:
        json.dump(status, f, ensure_ascii=False, indent=2)
    log.info(f"Saved diamond_status.json → {DIAMOND_STATUS_JSON}")


def load_vendors_from_env():
    vendor_defaults = [
        {"name": "loose-grown", "env_csv": "LOOSE_GROWN_CSV",  "env_cert": "LOOSE_GROWN_CERT_COL",  "env_price": "LOOSE_GROWN_PRICE_COL",  "env_disc": "LOOSE_GROWN_DISCOUNT",  "default_csv": "loosegrowndiamond.csv",    "default_cert": "sku",                "default_price": "price",            "default_disc": "0.70"},
        {"name": "brilliance",  "env_csv": "BRILLIANCE_CSV",   "env_cert": "BRILLIANCE_CERT_COL",   "env_price": "BRILLIANCE_PRICE_COL",   "env_disc": "BRILLIANCE_DISCOUNT",   "default_csv": "brilliance_diamonds.csv", "default_cert": "reportNumber",       "default_price": "price",            "default_disc": "0.70"},
        {"name": "luvansh",     "env_csv": "LUVANSH_CSV",      "env_cert": "LUVANSH_CERT_COL",      "env_price": "LUVANSH_PRICE_COL",      "env_disc": "LUVANSH_DISCOUNT",      "default_csv": "luvansh_diamonds.csv",    "default_cert": "certificate_number", "default_price": "discounted_price", "default_disc": "1.00"},
    ]

    vendors = []
    for v in vendor_defaults:
        csv_name  = get_env(v["env_csv"],   v["default_csv"])
        cert_col  = get_env(v["env_cert"],  v["default_cert"])
        price_col = get_env(v["env_price"], v["default_price"])
        discount  = float(get_env(v["env_disc"], v["default_disc"]))
        vendors.append({
            "name": v["name"], "csv": os.path.join(DIAMOND_FILES_DIR, csv_name),
            "cert_col": cert_col, "price_col": price_col, "discount": discount,
        })

    PC_csv_name = get_env("PC_CSV", "precious_carbon.csv")
    PC_csv      = os.path.join(DIAMOND_FILES_DIR, PC_csv_name)
    PC_discount = float(get_env("PC_DISCOUNT", "0.70"))
    return vendors, PC_csv, PC_discount


def build_compare_files(loaded: dict, discounts: dict) -> pd.DataFrame:
    log.info("=" * 50)
    log.info("STEP 1 — Building compare files")
    log.info(f"Vendors loaded: {list(loaded.keys())}")

    if "PC" not in loaded:
        log.warning("PC not loaded — falling back to outer join")
        combined = None
        for name, df in loaded.items():
            combined = df if combined is None else pd.merge(combined, df, on="certificate_number", how="outer")
    else:
        combined = loaded["PC"].copy()
        log.info(f"PC master list: {len(combined):,} cert IDs")
        for name, df in loaded.items():
            if name == "PC":
                continue
            combined = pd.merge(combined, df, on="certificate_number", how="left")
            matched = combined[f"{name} Price USD"].notna().sum()
            log.info(f"  '{name}': {matched:,} matched PC")

    if combined is None or combined.empty:
        log.error("No data loaded — aborting.")
        return pd.DataFrame()

    combined = combined.reset_index(drop=True)
    price_cols = [f"{name} Price USD" for name in loaded.keys() if f"{name} Price USD" in combined.columns]
    combined["_vendor_count"] = combined[price_cols].notna().sum(axis=1)

    CORE_VENDORS = ["loose-grown", "brilliance", "PC"]
    core_present = [n for n in CORE_VENDORS if f"{n} Price USD" in combined.columns]
    combined["_all_matched"]  = combined[[f"{n} Price USD" for n in core_present]].notna().all(axis=1) if core_present else combined["_vendor_count"] >= 2
    combined["_some_matched"] = combined["_vendor_count"] >= 2

    common_count  = combined["_all_matched"].sum()
    partial_count = combined["_some_matched"].sum()
    log.info(f"  Common (LG+BR+PC): {common_count:,}  |  Partial (2+): {partial_count:,}  |  All PC: {len(combined):,}")

    # Discounted price columns
    for name, disc in discounts.items():
        price_col = f"{name} Price USD"
        label     = f"-{round((1 - disc) * 100)}%"
        disc_col  = f"{name} {label} USD"
        if price_col in combined.columns:
            combined[price_col] = pd.to_numeric(combined[price_col], errors="coerce").round(0)
            combined[disc_col]  = (combined[price_col] * disc).round(0)

    # % diff vs PC
    if "PC" in discounts and "PC Price USD" in combined.columns:
        PC_disc     = discounts["PC"]
        PC_label    = f"-{round((1 - PC_disc) * 100)}%"
        PC_disc_col = f"PC {PC_label} USD"
        if PC_disc_col in combined.columns:
            for name in discounts:
                if name == "PC":
                    continue
                disc     = discounts[name]
                label    = f"-{round((1 - disc) * 100)}%"
                disc_col = f"{name} {label} USD"
                pct_col  = f"{name} vs PC %"
                if disc_col not in combined.columns:
                    continue
                vendor_price = pd.to_numeric(combined[disc_col], errors="coerce").replace(0, float("nan"))
                PC_price     = pd.to_numeric(combined[PC_disc_col], errors="coerce").replace(0, float("nan"))
                combined[pct_col] = ((vendor_price - PC_price) / PC_price * 100).round(2)
                log.info(f"Added % diff: {pct_col}  ({combined[pct_col].notna().sum():,} valid rows)")

            # ── PC vs vendor $ diff columns ───────────────────────
            # LG vs PC $  (LG price - PC -30%)
            lg_price_col = "loose-grown Price USD"
            if lg_price_col in combined.columns:
                pc_vals = pd.to_numeric(combined[PC_disc_col], errors="coerce").replace(0, float("nan"))
                lg_vals = pd.to_numeric(combined[lg_price_col], errors="coerce").replace(0, float("nan"))
                combined["loose-grown vs PC USD"] = (lg_vals - pc_vals).round(0)
                log.info(f"Added loose-grown vs PC USD  ({combined['loose-grown vs PC USD'].notna().sum():,} valid rows)")

            # Brilliance vs PC $  (BR -30% - PC -30%)
            br_disc     = discounts.get("brilliance", 0.70)
            br_label    = f"-{round((1 - br_disc) * 100)}%"
            br_disc_col = f"brilliance {br_label} USD"
            if br_disc_col in combined.columns:
                pc_vals = pd.to_numeric(combined[PC_disc_col], errors="coerce").replace(0, float("nan"))
                br_vals = pd.to_numeric(combined[br_disc_col], errors="coerce").replace(0, float("nan"))
                combined["brilliance vs PC USD"] = (br_vals - pc_vals).round(0)
                log.info(f"Added brilliance vs PC USD  ({combined['brilliance vs PC USD'].notna().sum():,} valid rows)")

    def vendor_list(row):
        return ", ".join(name for name in loaded.keys() if pd.notna(row.get(f"{name} Price USD")))
    combined["vendors_matched"] = combined.apply(vendor_list, axis=1)

    # Column ordering
    attr_cols = [c for c in list(PC_EXTRA_COLS.values()) if c in combined.columns]
    base_cols = ["certificate_number", "vendors_matched"] + attr_cols
    VENDOR_ORDER = ["loose-grown", "brilliance", "luvansh", "PC"]
    extra_vendors = [n for n in loaded.keys() if n not in VENDOR_ORDER and n != "PC"]
    ordered_vendors = ([n for n in VENDOR_ORDER if n in loaded and n != "PC"] + extra_vendors + (["PC"] if "PC" in loaded else []))
    vendor_cols = []
    for name in ordered_vendors:
        disc  = discounts.get(name, 0.70)
        label = f"-{round((1 - disc) * 100)}%"
        # Price → discounted price → $ vs PC → % vs PC
        for col in [
            f"{name} Price USD",
            f"{name} {label} USD",
            f"{name} vs PC USD",
            f"{name} vs PC %",
        ]:
            if col in combined.columns:
                vendor_cols.append(col)
    other_cols = [c for c in combined.columns if c not in base_cols + vendor_cols]
    combined   = combined[base_cols + vendor_cols + other_cols]

    def to_json_records(df: pd.DataFrame) -> list:
        import math
        out = df.drop(columns=["_all_matched", "_some_matched", "_vendor_count"], errors="ignore").copy()
        p_cols = [c for c in out.columns if "Price USD" in c or c.endswith("USD")]
        for col in p_cols:
            out[col] = pd.to_numeric(out[col], errors="coerce")
            out[col] = out[col].where(out[col].notna() & (out[col] != 0), other=None)
        pct_cols = [c for c in out.columns if c.endswith("%")]
        for col in pct_cols:
            out[col] = pd.to_numeric(out[col], errors="coerce")
            out[col] = out[col].where(out[col].notna(), other=None)
        records = out.where(out.notna(), other=None).to_dict(orient="records")
        clean = []
        for row in records:
            clean_row = {}
            for k, v in row.items():
                if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                    clean_row[k] = None
                else:
                    clean_row[k] = v
            clean.append(clean_row)
        return clean

    df_common  = combined[combined["_all_matched"]].copy()
    df_partial = combined[combined["_some_matched"]].copy()
    df_all     = combined.copy()

    common_records  = to_json_records(df_common)
    partial_records = to_json_records(df_partial)
    all_records     = to_json_records(df_all)

    with open(COMPARE_JSON, "w", encoding="utf-8") as f:
        json.dump(common_records, f, ensure_ascii=False, indent=2)
    with open(COMPARE_PARTIAL_JSON, "w", encoding="utf-8") as f:
        json.dump(partial_records, f, ensure_ascii=False, indent=2)
    with open(COMPARE_ALL_JSON, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    df_common.drop(columns=["_all_matched", "_some_matched", "_vendor_count"], errors="ignore").to_csv(COMPARE_CSV, index=False)

    log.info(f"compare.json ({len(common_records):,}) | compare_partial.json ({len(partial_records):,}) | compare_all.json ({len(all_records):,})")
    return combined, common_records, partial_records, all_records


# ══════════════════════════════════════════════════════════════
# STEP 2 — EXCEL COMPARISON SHEETS
# ══════════════════════════════════════════════════════════════

def categorize_diff(val):
    if pd.isna(val): return "N/A"
    val = abs(val)
    if val <= 100:    return "0-100"
    elif val <= 200:  return "100-200"
    elif val <= 300:  return "200-300"
    elif val <= 400:  return "300-400"
    elif val <= 500:  return "400-500"
    elif val <= 1000: return "500-1000"
    else:             return "1000+"


CATEGORY_COLORS = {
    "0-100": "C6EFCE", "100-200": "FFEB9C", "200-300": "FFE066",
    "300-400": "FFCC99", "400-500": "FFA07A",
    "500-1000": "FF9999", "1000+": "CC0000", "N/A": "D9D9D9",
}
DARK_CATEGORIES = {"1000+"}


def _write_sheet_with_formatting(df, writer, sheet_name, highlight_col=None, category_col=None):
    try:
        def highlight_loss(v):
            return ["background-color: #FF9999" if x < 0 else "" for x in v]
        if highlight_col and highlight_col in df.columns:
            styled = df.style.apply(highlight_loss, subset=[highlight_col])
            styled.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except AttributeError:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws      = writer.sheets[sheet_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    if category_col and category_col in headers:
        cat_col_idx = headers.index(category_col) + 1
        for row in range(2, len(df) + 2):
            cell  = ws.cell(row=row, column=cat_col_idx)
            cat   = cell.value
            color = CATEGORY_COLORS.get(str(cat), "FFFFFF")
            cell.fill = PatternFill("solid", start_color=color, fgColor=color)
            cell.font = Font(color="FFFFFF" if cat in DARK_CATEGORIES else "000000", bold=True)
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def compare_two_from_combined(combined: pd.DataFrame, name_a, discount_a, name_b, discount_b) -> pd.DataFrame:
    label_a    = f"-{round((1 - discount_a) * 100)}%"
    label_b    = f"-{round((1 - discount_b) * 100)}%"
    price_a    = f"{name_a} Price USD"
    price_b    = f"{name_b} Price USD"
    disc_col_a = f"{name_a} {label_a} USD"
    disc_col_b = f"{name_b} {label_b} USD"
    needed = ["certificate_number", price_a, disc_col_a, price_b, disc_col_b]
    avail  = [c for c in needed if c in combined.columns]
    df     = combined[avail].dropna(subset=[price_a, price_b]).copy()
    if df.empty:
        return pd.DataFrame()
    compare_col = f"Compare ({name_b}{label_b} - {name_a}{label_a}) USD"
    if disc_col_a in df.columns and disc_col_b in df.columns:
        df[compare_col] = (df[disc_col_b] - df[disc_col_a]).round(2)
    return df.reset_index(drop=True)


def compare_all_three_from_combined(combined: pd.DataFrame, discounts: dict, names: list, writer):
    price_cols = [f"{n} Price USD" for n in names]
    df = combined.dropna(subset=price_cols).copy()
    if df.empty:
        return False
    disc_cols = {}
    for name in names:
        disc     = discounts[name]
        label    = f"-{round((1 - disc) * 100)}%"
        disc_col = f"{name} {label} USD"
        disc_cols[name] = disc_col
        if disc_col not in df.columns:
            df[disc_col] = (pd.to_numeric(df[f"{name} Price USD"], errors="coerce") * disc).round(0)
    n = names
    diff_ab = f"{n[0]} vs {n[1]} Diff USD"
    diff_bc = f"{n[1]} vs {n[2]} Diff USD"
    diff_ca = f"{n[2]} vs {n[0]} Diff USD"
    df[diff_ab]        = (df[disc_cols[n[0]]] - df[disc_cols[n[1]]]).abs().round(2)
    df[diff_bc]        = (df[disc_cols[n[1]]] - df[disc_cols[n[2]]]).abs().round(2)
    df[diff_ca]        = (df[disc_cols[n[2]]] - df[disc_cols[n[0]]]).abs().round(2)
    df["Min Diff USD"] = df[[diff_ab, diff_bc, diff_ca]].min(axis=1).round(2)
    df["Category"]     = df["Min Diff USD"].apply(categorize_diff)
    sheet_name = " vs ".join(names)[:31]
    _write_sheet_with_formatting(df, writer, sheet_name, category_col="Category")
    return True


def compare_all_vendors(vendors, PC_csv=None, PC_discount=0.70, default_discount=0.70):
    log.info("=" * 50)
    log.info("Starting vendor comparison pipeline")

    loaded    = {}
    discounts = {}
    for v in vendors:
        loaded[v["name"]]    = load_vendor(v["csv"], v["cert_col"], v["price_col"], v["name"])
        discounts[v["name"]] = v.get("discount", default_discount)
    if PC_csv:
        loaded["PC"]    = load_PC(PC_csv)
        discounts["PC"] = PC_discount

    # ── STEP 1: build compare files ───────────────────────────
    result = build_compare_files(loaded, discounts)
    if isinstance(result, tuple):
        combined, common_records, partial_records, all_records = result
    else:
        combined = result
        common_records = partial_records = all_records = []

    if combined is None or (hasattr(combined, "empty") and combined.empty):
        log.error("compare.csv is empty — aborting.")
        return {}

    # ── NEW: build BR vs LG file ──────────────────────────────
    br_vs_lg_count = build_br_vs_lg(loaded, discounts)

    # ── Save status ───────────────────────────────────────────
    _save_diamond_status(combined, loaded, discounts, common_records, partial_records, all_records, br_vs_lg_count)


# ══════════════════════════════════════════════════════════════
# FILE AGE CHECK
# ══════════════════════════════════════════════════════════════

def get_env(key, default):
    return os.getenv(key, default)


def check_file_age(filepath: str, label: str) -> dict:
    if not os.path.exists(filepath):
        log.error(f"[{label}] MISSING: {filepath}")
        return {"file": os.path.basename(filepath), "vendor": label,
                "status": "MISSING", "last_modified": None, "age_days": None, "ok": False}
    mtime    = os.path.getmtime(filepath)
    mod_dt   = datetime.datetime.fromtimestamp(mtime)
    age_days = (datetime.datetime.now() - mod_dt).days
    is_fresh = age_days < MAX_AGE_DAYS
    status   = "OK" if is_fresh else "STALE"
    if is_fresh:
        log.info(f"[{label}] {status} — {os.path.basename(filepath)} ({age_days}d old)")
    else:
        log.warning(f"[{label}] {status} — {os.path.basename(filepath)} ({age_days}d old, last: {mod_dt.strftime('%Y-%m-%d %H:%M:%S')})")
    return {
        "file": os.path.basename(filepath), "vendor": label,
        "status": status, "last_modified": mod_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "age_days": age_days, "ok": is_fresh,
    }


def check_all_files_and_run():
    log.info("=" * 50)
    log.info("FILE AGE CHECK")
    vendors, PC_csv, PC_discount = load_vendors_from_env()
    all_files = [(v["csv"], v["name"]) for v in vendors] + [(PC_csv, "PC")]
    statuses  = []
    stale_warn = []
    for fpath, label in all_files:
        status = check_file_age(fpath, label)
        statuses.append(status)
        if not status["ok"] and status["status"] == "STALE":
            stale_warn.append(status)
    status_report = {
        "checked_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "all_fresh":  all(s["ok"] for s in statuses),
        "files":      statuses,
    }
    os.makedirs(DIAMOND_FILES_DIR, exist_ok=True)
    with open(FILE_STATUS_JSON, "w") as f:
        json.dump(status_report, f, indent=2)
    missing = [s for s in statuses if s["status"] == "MISSING"]
    if missing:
        log.error(f"ABORT: missing files: {[m['file'] for m in missing]}")
        return None
    return compare_all_vendors(vendors=vendors, PC_csv=PC_csv, PC_discount=PC_discount)


if __name__ == "__main__":
    check_all_files_and_run()